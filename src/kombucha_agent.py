"""
=============================================================================
  DYNAMIC PRODUCTION PLANNING AND CONTROL AI AGENT
  Kombucha Brewery - Industrial Engineering Portfolio Project
  Author: Production Planning AI System
  Description: Reads Excel data, plans brewing batches, allocates tanks,
               tracks fermentation, schedules bottling, detects bottlenecks,
               and writes production plans back to Excel.
=============================================================================
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────────
EXCEL_FILE         = "kombucha_brewery.xlsx"
LITERS_PER_BOTTLE  = 0.355          # 12 oz bottle
FERMENTATION_DAYS  = 10             # default fermentation window
PLANNING_HORIZON   = 21             # days to plan ahead
TODAY              = datetime.today().date()

# Colour palette (ARGB hex, no #)
CLR_HEADER_DARK    = "FF1B3A5C"     # deep navy
CLR_HEADER_MID     = "FF2E75B6"     # corporate blue
CLR_HEADER_LIGHT   = "FFD6E4F0"     # pale blue
CLR_ACCENT_GREEN   = "FF375623"     # dark green
CLR_ACCENT_ORANGE  = "FFBF5700"     # amber
CLR_ALERT_RED      = "FFC00000"     # alert red
CLR_ALERT_YELLOW   = "FFFFF2CC"     # warning yellow bg
CLR_ROW_ALT        = "FFF2F7FB"     # alternate row
CLR_WHITE          = "FFFFFFFF"
CLR_BORDER         = "FFB8CCE4"

# ─────────────────────────────────────────────────────────────────
#  STEP 1 – LOAD ALL EXCEL DATA
# ─────────────────────────────────────────────────────────────────
def load_data(filepath: str) -> dict:
    """Load all sheets from the brewery Excel workbook."""
    print("\n📂  STEP 1 – Loading brewery data from Excel ...")
    # Row 0 = decorative title, row 1 = actual column headers → header=1
    sheets = pd.read_excel(filepath, sheet_name=None, header=1)
    required = [
        "Orders", "Inventory", "Recipe_BOM",
        "Fermentation_Tanks", "Bottling_Line",
        "Production_Plan", "Actual_Production"
    ]
    for s in required:
        if s not in sheets:
            raise ValueError(f"Missing required sheet: '{s}'")

    # Parse dates
    for col in ["Due_Date"]:
        if col in sheets["Orders"].columns:
            sheets["Orders"][col] = pd.to_datetime(
                sheets["Orders"][col], errors="coerce"
            ).dt.date

    for col in ["Available_Date"]:
        if col in sheets["Fermentation_Tanks"].columns:
            sheets["Fermentation_Tanks"][col] = pd.to_datetime(
                sheets["Fermentation_Tanks"][col], errors="coerce"
            ).dt.date

    print(f"  ✔  Loaded {len(sheets['Orders'])} orders")
    print(f"  ✔  Loaded {len(sheets['Inventory'])} inventory items")
    print(f"  ✔  Loaded {len(sheets['Fermentation_Tanks'])} fermentation tanks")
    print(f"  ✔  Loaded {len(sheets['Bottling_Line'])} bottling lines")
    return sheets


# ─────────────────────────────────────────────────────────────────
#  STEP 2 – FORECAST DEMAND
# ─────────────────────────────────────────────────────────────────
def forecast_demand(orders_df: pd.DataFrame) -> pd.DataFrame:
    """
    Summarise demand per flavor, sorted by priority and due date.
    Priority 1 = highest urgency.
    """
    print("\n📊  STEP 2 – Forecasting production demand ...")
    df = orders_df.copy()
    df["Due_Date"] = pd.to_datetime(df["Due_Date"], errors="coerce").dt.date
    df = df.sort_values(["Priority", "Due_Date"], ascending=[True, True])

    summary = (
        df.groupby("Flavor")
          .agg(
              Total_Bottles=("Quantity_Bottles", "sum"),
              Earliest_Due=("Due_Date", "min"),
              Orders_Count=("Order_ID", "count"),
              Max_Priority=("Priority", "min")  # lower number = higher priority
          )
          .reset_index()
          .sort_values(["Max_Priority", "Earliest_Due"])
    )
    print(f"  ✔  {len(summary)} flavors in demand forecast")
    for _, row in summary.iterrows():
        print(f"     • {row['Flavor']:25s} → {int(row['Total_Bottles']):>6,} bottles  "
              f"(earliest due: {row['Earliest_Due']})")
    return summary


# ─────────────────────────────────────────────────────────────────
#  STEP 3 – CONVERT BOTTLES TO BATCH SIZES
# ─────────────────────────────────────────────────────────────────
def convert_to_batches(demand_df: pd.DataFrame, tanks_df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert bottle demand to liters, then split into feasible batch sizes
    based on the largest available tank capacity.
    """
    print("\n⚗️   STEP 3 – Converting bottle demand to batch sizes ...")
    max_tank_cap = tanks_df["Capacity_Liters"].max()

    batches = []
    for _, row in demand_df.iterrows():
        total_liters = row["Total_Bottles"] * LITERS_PER_BOTTLE
        n_full_batches = int(total_liters // max_tank_cap)
        remainder = total_liters % max_tank_cap

        for i in range(n_full_batches):
            batches.append({
                "Flavor":          row["Flavor"],
                "Batch_Liters":    max_tank_cap,
                "Bottles_Needed":  int(max_tank_cap / LITERS_PER_BOTTLE),
                "Due_Date":        row["Earliest_Due"],
                "Priority":        row["Max_Priority"],
                "Batch_Seq":       i + 1
            })
        if remainder > 0:
            batches.append({
                "Flavor":          row["Flavor"],
                "Batch_Liters":    round(remainder, 1),
                "Bottles_Needed":  int(remainder / LITERS_PER_BOTTLE),
                "Due_Date":        row["Earliest_Due"],
                "Priority":        row["Max_Priority"],
                "Batch_Seq":       n_full_batches + 1
            })

    result = pd.DataFrame(batches)
    print(f"  ✔  {len(result)} production batches planned "
          f"(max tank: {max_tank_cap}L)")
    return result


# ─────────────────────────────────────────────────────────────────
#  STEP 4 – CHECK RAW MATERIAL AVAILABILITY
# ─────────────────────────────────────────────────────────────────
def check_materials(batches_df: pd.DataFrame,
                    inventory_df: pd.DataFrame,
                    bom_df: pd.DataFrame) -> list:
    """
    Check if current inventory can cover all planned batches.
    Returns list of shortage alert dicts.
    """
    print("\n🧪  STEP 4 – Checking raw material availability ...")
    alerts = []
    inv = inventory_df.set_index("Material_Name").to_dict(orient="index")

    for flavor in batches_df["Flavor"].unique():
        flavor_batches = batches_df[batches_df["Flavor"] == flavor]
        total_liters   = flavor_batches["Batch_Liters"].sum()
        n_batches      = len(flavor_batches)

        recipe = bom_df[bom_df["Flavor"].str.strip() == flavor.strip()]
        if recipe.empty:
            print(f"  ⚠️   No BOM found for '{flavor}' – skipping material check")
            continue

        for _, ing in recipe.iterrows():
            material   = ing["Ingredient"]
            qty_needed = ing["Quantity_Per_Batch"] * n_batches

            if material not in inv:
                alerts.append({
                    "Type": "MISSING_MATERIAL",
                    "Material": material,
                    "Flavor": flavor,
                    "Qty_Needed": qty_needed,
                    "Stock": 0,
                    "Shortage": qty_needed,
                    "Message": (
                        f"⛔  '{material}' not found in inventory "
                        f"(needed for {flavor})"
                    )
                })
                continue

            stock    = inv[material].get("Stock_Level", 0)
            reorder  = inv[material].get("Reorder_Point", 0)
            safety   = inv[material].get("Safety_Stock", 0)
            shortage = max(0, qty_needed - stock)
            unit     = inv[material].get("Unit", "units")

            if shortage > 0:
                alerts.append({
                    "Type": "SHORTAGE",
                    "Material": material,
                    "Flavor": flavor,
                    "Qty_Needed": round(qty_needed, 2),
                    "Stock": stock,
                    "Shortage": round(shortage, 2),
                    "Message": (
                        f"🔴  SHORTAGE: '{material}' needs {qty_needed:.1f} {unit} "
                        f"but only {stock} in stock  (short by {shortage:.1f})"
                    )
                })
            elif stock <= reorder:
                alerts.append({
                    "Type": "LOW_STOCK",
                    "Material": material,
                    "Flavor": flavor,
                    "Qty_Needed": round(qty_needed, 2),
                    "Stock": stock,
                    "Shortage": 0,
                    "Message": (
                        f"🟡  LOW STOCK: '{material}' is at {stock} {unit} "
                        f"(reorder point: {reorder}) – place order soon"
                    )
                })

    if not alerts:
        print("  ✔  All materials sufficient for planned batches")
    else:
        for a in alerts:
            print(f"  {a['Message']}")
    return alerts


# ─────────────────────────────────────────────────────────────────
#  STEP 5 & 6 – ASSIGN TANKS AND SCHEDULE FERMENTATION
# ─────────────────────────────────────────────────────────────────
def assign_tanks_and_schedule(batches_df: pd.DataFrame,
                               tanks_df: pd.DataFrame) -> pd.DataFrame:
    """
    Greedy tank assignment: for each batch (highest priority first),
    pick the earliest available tank that fits the batch volume.
    Returns production schedule with fermentation dates.
    """
    print("\n🏭  STEP 5 & 6 – Assigning tanks and scheduling fermentation ...")

    # Maintain a mutable copy of tank availability
    tank_state = tanks_df.copy()
    tank_state["Available_Date"] = pd.to_datetime(
        tank_state["Available_Date"], errors="coerce"
    ).dt.date

    schedule = []
    batches_sorted = batches_df.sort_values(["Priority", "Due_Date"])

    for _, batch in batches_sorted.iterrows():
        needed_liters = batch["Batch_Liters"]

        # Filter tanks with enough capacity
        eligible = tank_state[
            tank_state["Capacity_Liters"] >= needed_liters
        ].copy()

        if eligible.empty:
            print(f"  ⚠️   No tank large enough for {batch['Flavor']} "
                  f"batch ({needed_liters}L) – BOTTLENECK DETECTED")
            schedule.append({
                "Flavor":             batch["Flavor"],
                "Batch_Size_Liters":  needed_liters,
                "Assigned_Tank":      "UNASSIGNED",
                "Fermentation_Start": None,
                "Fermentation_End":   None,
                "Bottling_Date":      None,
                "Bottles_Planned":    batch["Bottles_Needed"],
                "Due_Date":           batch["Due_Date"],
                "Priority":           batch["Priority"],
                "Status":             "NO_TANK_AVAILABLE"
            })
            continue

        # Pick the tank that becomes available soonest
        eligible = eligible.sort_values("Available_Date")
        chosen = eligible.iloc[0]
        tank_id = chosen["Tank_ID"]

        ferm_start = max(TODAY, chosen["Available_Date"])
        ferm_end   = ferm_start + timedelta(days=FERMENTATION_DAYS)
        bottle_date = ferm_end + timedelta(days=1)  # bottling day after fermentation

        # Determine on-time status
        due = batch["Due_Date"]
        late = bottle_date > due if due else False
        status = "AT_RISK_LATE" if late else "ON_SCHEDULE"

        schedule.append({
            "Flavor":             batch["Flavor"],
            "Batch_Size_Liters":  needed_liters,
            "Assigned_Tank":      tank_id,
            "Fermentation_Start": ferm_start,
            "Fermentation_End":   ferm_end,
            "Bottling_Date":      bottle_date,
            "Bottles_Planned":    batch["Bottles_Needed"],
            "Due_Date":           due,
            "Priority":           batch["Priority"],
            "Status":             status
        })

        # Update tank availability (tank freed after fermentation ends)
        tank_state.loc[tank_state["Tank_ID"] == tank_id,
                       "Available_Date"] = ferm_end

    result = pd.DataFrame(schedule)
    on_sched = (result["Status"] == "ON_SCHEDULE").sum()
    at_risk   = (result["Status"] == "AT_RISK_LATE").sum()
    unassigned = (result["Status"] == "NO_TANK_AVAILABLE").sum()
    print(f"  ✔  {len(result)} batches scheduled: "
          f"{on_sched} on-time | {at_risk} at-risk | {unassigned} unassigned")
    return result


# ─────────────────────────────────────────────────────────────────
#  STEP 7 – SCHEDULE BOTTLING OPERATIONS
# ─────────────────────────────────────────────────────────────────
def schedule_bottling(schedule_df: pd.DataFrame,
                      bottling_df: pd.DataFrame) -> pd.DataFrame:
    """
    Assign bottling slots considering line capacity (bottles/day).
    Detects overload days and records actual scheduled bottling date.
    """
    print("\n🍾  STEP 7 – Scheduling bottling operations ...")

    total_capacity_per_day = int(
        (bottling_df["Bottles_Per_Hour"] * bottling_df["Operating_Hours_Per_Day"]).sum()
    )
    print(f"  ✔  Total bottling capacity: {total_capacity_per_day:,} bottles/day")

    # Build daily capacity tracker
    daily_usage: dict = {}

    result = schedule_df.copy()
    result["Actual_Bottling_Date"] = None
    result["Bottling_Overload"]    = False

    for idx, row in result.iterrows():
        if row["Bottling_Date"] is None:
            continue

        bottles_needed = row["Bottles_Planned"]
        planned_date   = row["Bottling_Date"]
        scheduled_date = planned_date

        # Roll forward if capacity exceeded
        max_search = 14
        for _ in range(max_search):
            used = daily_usage.get(scheduled_date, 0)
            if used + bottles_needed <= total_capacity_per_day:
                break
            scheduled_date += timedelta(days=1)

        daily_usage[scheduled_date] = daily_usage.get(scheduled_date, 0) + bottles_needed
        overload = daily_usage[scheduled_date] > total_capacity_per_day

        result.at[idx, "Actual_Bottling_Date"] = scheduled_date
        result.at[idx, "Bottling_Overload"]    = overload

    # Report overloaded days
    overloaded_days = {
        d: v for d, v in daily_usage.items()
        if v > total_capacity_per_day
    }
    if overloaded_days:
        for d, v in overloaded_days.items():
            print(f"  ⚠️   Bottling OVERLOAD on {d}: "
                  f"{v:,} bottles vs {total_capacity_per_day:,} capacity")
    else:
        print("  ✔  No bottling capacity overloads detected")

    return result


# ─────────────────────────────────────────────────────────────────
#  STEP 8 – DETECT BOTTLENECKS
# ─────────────────────────────────────────────────────────────────
def detect_bottlenecks(schedule_df: pd.DataFrame,
                        material_alerts: list,
                        tanks_df: pd.DataFrame,
                        bottling_df: pd.DataFrame) -> list:
    """
    Consolidate all bottleneck signals into actionable insights.
    """
    print("\n🔍  STEP 8 – Detecting bottlenecks and generating insights ...")
    insights = []

    # ── Tank bottlenecks
    for _, row in schedule_df.iterrows():
        if row["Status"] == "NO_TANK_AVAILABLE":
            insights.append({
                "Category": "TANK_SHORTAGE",
                "Severity": "HIGH",
                "Message":  (
                    f"Tank shortage: {row['Flavor']} batch of "
                    f"{row['Batch_Size_Liters']}L cannot be assigned – "
                    f"consider adding fermentation capacity."
                )
            })

    # ── Tank availability timeline
    for _, tank in tanks_df.iterrows():
        avail = tank["Available_Date"]
        if isinstance(avail, datetime):
            avail = avail.date()
        if avail and avail > TODAY:
            insights.append({
                "Category": "TANK_STATUS",
                "Severity": "INFO",
                "Message":  (
                    f"Tank {tank['Tank_ID']} will become available on {avail}."
                )
            })

    # ── Ingredient shortages
    for alert in material_alerts:
        if alert["Type"] == "SHORTAGE":
            insights.append({
                "Category": "INGREDIENT_SHORTAGE",
                "Severity": "HIGH",
                "Message":  alert["Message"]
            })
        elif alert["Type"] == "LOW_STOCK":
            insights.append({
                "Category": "LOW_INVENTORY",
                "Severity": "MEDIUM",
                "Message":  alert["Message"]
            })

    # ── Bottling capacity
    overloads = schedule_df[schedule_df.get("Bottling_Overload", False) == True]
    if not overloads.empty:
        dates = overloads["Actual_Bottling_Date"].unique()
        for d in dates:
            insights.append({
                "Category": "BOTTLING_OVERLOAD",
                "Severity": "HIGH",
                "Message":  f"Bottling line capacity EXCEEDED on {d}."
            })

    # ── Late orders
    late_orders = schedule_df[schedule_df["Status"] == "AT_RISK_LATE"]
    if not late_orders.empty:
        for _, row in late_orders.iterrows():
            insights.append({
                "Category": "LATE_ORDER_RISK",
                "Severity": "HIGH",
                "Message":  (
                    f"⚠️  {row['Flavor']} batch (due {row['Due_Date']}) "
                    f"will bottle on {row.get('Actual_Bottling_Date', row['Bottling_Date'])} – "
                    f"AT RISK of being late."
                )
            })

    # ── Start-brew recommendations
    start_soon = schedule_df[
        pd.to_datetime(schedule_df["Fermentation_Start"], errors="coerce").dt.date
        == TODAY + timedelta(days=1)
    ]
    for _, row in start_soon.iterrows():
        insights.append({
            "Category": "ACTION_REQUIRED",
            "Severity": "INFO",
            "Message":  (
                f"▶  Start brewing {row['Flavor']} tomorrow "
                f"({row['Fermentation_Start']}) in tank {row['Assigned_Tank']} "
                f"to meet demand."
            )
        })

    # Print all insights
    for ins in insights:
        sev_icon = {"HIGH": "🔴", "MEDIUM": "🟡", "INFO": "🔵"}.get(ins["Severity"], "⚪")
        print(f"  {sev_icon}  [{ins['Category']}] {ins['Message']}")

    print(f"\n  ✔  {len(insights)} insights generated")
    return insights


# ─────────────────────────────────────────────────────────────────
#  HELPER – CELL STYLING
# ─────────────────────────────────────────────────────────────────
def _thin_border() -> Border:
    side = Side(style="thin", color=CLR_BORDER[2:])  # strip FF prefix
    return Border(left=side, right=side, top=side, bottom=side)

def _header_style(cell, bg: str = CLR_HEADER_DARK, font_size: int = 10):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=font_size)
    cell.fill      = PatternFill("solid", fgColor=bg[2:])
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border    = _thin_border()

def _data_style(cell, alt: bool = False):
    cell.fill      = PatternFill("solid",
                                  fgColor=(CLR_ROW_ALT if alt else CLR_WHITE)[2:])
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center")
    cell.border    = _thin_border()

def _status_color(cell, status: str):
    colors = {
        "ON_SCHEDULE":       "D9EAD3",
        "AT_RISK_LATE":      "FFE599",
        "NO_TANK_AVAILABLE": "F4CCCC"
    }
    cell.fill = PatternFill("solid", fgColor=colors.get(status, "FFFFFF"))

def _autofit_columns(ws, min_width: int = 10, max_width: int = 40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len  = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width,
                                                       min(max_width, max_len + 3))


# ─────────────────────────────────────────────────────────────────
#  STEP 9 – WRITE PRODUCTION SCHEDULE BACK TO EXCEL
# ─────────────────────────────────────────────────────────────────
def write_to_excel(filepath: str,
                   schedule_df: pd.DataFrame,
                   insights: list,
                   material_alerts: list):
    """
    Write the production schedule, bottleneck report, and summary dashboard
    back into the Excel workbook using professional formatting.
    """
    print("\n💾  STEP 9 – Writing production schedule to Excel ...")
    wb = load_workbook(filepath)

    # ─── 9a. Production_Plan sheet ────────────────────────────────
    ws_plan = wb["Production_Plan"]
    ws_plan.delete_rows(2, ws_plan.max_row)  # clear existing data rows

    headers = [
        "Date", "Flavor", "Batch_Size_Liters", "Assigned_Tank",
        "Fermentation_Start", "Fermentation_End",
        "Bottling_Date", "Actual_Bottling_Date",
        "Bottles_Planned", "Due_Date", "Priority", "Status"
    ]

    # Re-write header row
    ws_plan.delete_rows(1, 1)
    ws_plan.insert_rows(1)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_plan.cell(row=1, column=col_idx, value=h)
        _header_style(cell, CLR_HEADER_DARK)
    ws_plan.row_dimensions[1].height = 30

    for row_idx, row in schedule_df.iterrows():
        r = row_idx + 2  # row 1 = header
        values = [
            str(TODAY),
            row.get("Flavor"),
            row.get("Batch_Size_Liters"),
            row.get("Assigned_Tank"),
            str(row.get("Fermentation_Start")) if row.get("Fermentation_Start") else "",
            str(row.get("Fermentation_End"))   if row.get("Fermentation_End") else "",
            str(row.get("Bottling_Date"))       if row.get("Bottling_Date") else "",
            str(row.get("Actual_Bottling_Date")) if row.get("Actual_Bottling_Date") else "",
            row.get("Bottles_Planned"),
            str(row.get("Due_Date")) if row.get("Due_Date") else "",
            row.get("Priority"),
            row.get("Status")
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_plan.cell(row=r, column=col_idx, value=val)
            alt  = (r % 2 == 0)
            _data_style(cell, alt=alt)

        # Color-code Status cell
        status_cell = ws_plan.cell(row=r, column=12)
        _status_color(status_cell, row.get("Status", ""))

    _autofit_columns(ws_plan)
    ws_plan.freeze_panes = "A2"

    # ─── 9b. Bottleneck_Report sheet ─────────────────────────────
    if "Bottleneck_Report" in wb.sheetnames:
        del wb["Bottleneck_Report"]
    ws_bn = wb.create_sheet("Bottleneck_Report")

    # Title
    ws_bn.merge_cells("A1:D1")
    title_cell = ws_bn["A1"]
    title_cell.value     = "🔍  BOTTLENECK & INSIGHTS REPORT"
    title_cell.font      = Font(name="Arial", bold=True, size=13,
                                color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor=CLR_HEADER_DARK[2:])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_bn.row_dimensions[1].height = 36

    bn_headers = ["#", "Category", "Severity", "Message"]
    for col_idx, h in enumerate(bn_headers, start=1):
        cell = ws_bn.cell(row=2, column=col_idx, value=h)
        _header_style(cell, CLR_HEADER_MID)

    sev_colors = {"HIGH": "F4CCCC", "MEDIUM": "FFE599", "INFO": "D9EAD3"}
    for i, ins in enumerate(insights, start=1):
        r = i + 2
        row_vals = [i, ins["Category"], ins["Severity"], ins["Message"]]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_bn.cell(row=r, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _thin_border()
            if col_idx == 3:
                sev = ins.get("Severity", "INFO")
                cell.fill = PatternFill("solid",
                                         fgColor=sev_colors.get(sev, "FFFFFF"))
        ws_bn.row_dimensions[r].height = 25

    ws_bn.column_dimensions["A"].width = 5
    ws_bn.column_dimensions["B"].width = 22
    ws_bn.column_dimensions["C"].width = 12
    ws_bn.column_dimensions["D"].width = 80
    ws_bn.freeze_panes = "A3"

    # ─── 9c. Ingredient_Alerts sheet ─────────────────────────────
    if "Ingredient_Alerts" in wb.sheetnames:
        del wb["Ingredient_Alerts"]
    ws_ia = wb.create_sheet("Ingredient_Alerts")

    ws_ia.merge_cells("A1:G1")
    t2 = ws_ia["A1"]
    t2.value     = "🧪  INGREDIENT SHORTAGE & LOW STOCK ALERTS"
    t2.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t2.fill      = PatternFill("solid", fgColor=CLR_ALERT_RED[2:])
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_ia.row_dimensions[1].height = 36

    ia_hdrs = ["#", "Type", "Material", "Flavor", "Qty_Needed", "Stock", "Shortage"]
    for col_idx, h in enumerate(ia_hdrs, start=1):
        cell = ws_ia.cell(row=2, column=col_idx, value=h)
        _header_style(cell, CLR_ALERT_RED)

    for i, alert in enumerate(material_alerts, start=1):
        r = i + 2
        row_vals = [
            i, alert["Type"], alert["Material"], alert["Flavor"],
            alert["Qty_Needed"], alert["Stock"], alert["Shortage"]
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_ia.cell(row=r, column=col_idx, value=val)
            _data_style(cell, alt=(r % 2 == 0))
        if alert["Type"] == "SHORTAGE":
            for c in range(1, 8):
                ws_ia.cell(row=r, column=c).fill = PatternFill(
                    "solid", fgColor="F4CCCC")

    _autofit_columns(ws_ia)

    # ─── 9d. Dashboard sheet ─────────────────────────────────────
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws_db = wb.create_sheet("Dashboard", 0)  # first tab
    _build_dashboard(ws_db, schedule_df, insights, material_alerts)

    wb.save(filepath)
    print(f"  ✔  Workbook saved: {filepath}")


def _build_dashboard(ws, schedule_df: pd.DataFrame,
                     insights: list, material_alerts: list):
    """Build a KPI summary dashboard sheet."""
    ws.sheet_view.showGridLines = False

    # Title bar
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = "🍵  KOMBUCHA BREWERY  |  PRODUCTION PLANNING DASHBOARD"
    t.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=CLR_HEADER_DARK[2:])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    sub_date = ws["A2"]
    ws.merge_cells("A2:H2")
    sub_date.value     = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    sub_date.font      = Font(name="Arial", italic=True, size=10, color="555555")
    sub_date.alignment = Alignment(horizontal="center")
    sub_date.fill      = PatternFill("solid", fgColor=CLR_HEADER_LIGHT[2:])

    # ── KPI Cards (row 4–6)
    kpis = [
        ("Total Batches",   len(schedule_df),                       CLR_HEADER_MID),
        ("On Schedule",     (schedule_df["Status"]=="ON_SCHEDULE").sum(),    "375623"),
        ("At Risk",         (schedule_df["Status"]=="AT_RISK_LATE").sum(),   "BF5700"),
        ("Unassigned",      (schedule_df["Status"]=="NO_TANK_AVAILABLE").sum(), "C00000"),
        ("Alerts",          len([a for a in insights if a["Severity"]=="HIGH"]), "C00000"),
        ("Low Stock Items", len([a for a in material_alerts if a["Type"]=="LOW_STOCK"]), "BF5700"),
    ]

    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[6].height = 8

    col_map = [1, 2, 3, 4, 5, 6]
    for col_offset, (label, value, color) in zip(col_map, kpis):
        label_cell = ws.cell(row=4, column=col_offset, value=label)
        label_cell.font      = Font(name="Arial", bold=True, size=9,
                                     color="FFFFFF")
        label_cell.fill      = PatternFill("solid", fgColor=color)
        label_cell.alignment = Alignment(horizontal="center", vertical="bottom")

        val_cell = ws.cell(row=5, column=col_offset, value=value)
        val_cell.font      = Font(name="Arial", bold=True, size=20,
                                   color=color)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.fill      = PatternFill("solid", fgColor="F8F8F8")

    # ── Production Schedule summary (row 8+)
    ws.row_dimensions[7].height = 12

    hdr_row = 8
    sched_hdrs = [
        "Flavor", "Batch (L)", "Tank", "Ferm. Start",
        "Ferm. End", "Bottling Date", "Bottles", "Status"
    ]
    for col_idx, h in enumerate(sched_hdrs, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=h)
        _header_style(cell, CLR_HEADER_MID, font_size=9)
    ws.row_dimensions[hdr_row].height = 24

    for row_idx, row in schedule_df.iterrows():
        r = hdr_row + 1 + row_idx
        row_vals = [
            row.get("Flavor"),
            row.get("Batch_Size_Liters"),
            row.get("Assigned_Tank"),
            str(row.get("Fermentation_Start")) if row.get("Fermentation_Start") else "",
            str(row.get("Fermentation_End"))   if row.get("Fermentation_End") else "",
            str(row.get("Actual_Bottling_Date") or row.get("Bottling_Date")) if row.get("Bottling_Date") else "",
            row.get("Bottles_Planned"),
            row.get("Status")
        ]
        alt = (r % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            _data_style(cell, alt=alt)
        status_cell = ws.cell(row=r, column=8)
        _status_color(status_cell, row.get("Status", ""))
        ws.row_dimensions[r].height = 18

    _autofit_columns(ws, min_width=12)
    ws.freeze_panes = "A9"


# ─────────────────────────────────────────────────────────────────
#  MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────
def run_agent(filepath: str = EXCEL_FILE):
    """
    Orchestrate all nine steps of the Production Planning AI Agent.
    """
    print("=" * 65)
    print("  KOMBUCHA BREWERY — DYNAMIC PRODUCTION PLANNING AI AGENT")
    print(f"  Planning Date: {TODAY}")
    print("=" * 65)

    # STEP 1 – Load
    data = load_data(filepath)

    orders_df   = data["Orders"]
    inv_df      = data["Inventory"]
    bom_df      = data["Recipe_BOM"]
    tanks_df    = data["Fermentation_Tanks"]
    bottling_df = data["Bottling_Line"]

    # STEP 2 – Forecast demand
    demand_df = forecast_demand(orders_df)

    # STEP 3 – Convert to batches
    batches_df = convert_to_batches(demand_df, tanks_df)

    # STEP 4 – Check materials
    material_alerts = check_materials(batches_df, inv_df, bom_df)

    # STEPS 5 & 6 – Assign tanks + schedule fermentation
    schedule_df = assign_tanks_and_schedule(batches_df, tanks_df)

    # STEP 7 – Schedule bottling
    schedule_df = schedule_bottling(schedule_df, bottling_df)

    # STEP 8 – Detect bottlenecks
    insights = detect_bottlenecks(schedule_df, material_alerts,
                                   tanks_df, bottling_df)

    # STEP 9 – Write back to Excel
    write_to_excel(filepath, schedule_df, insights, material_alerts)

    # ── Final summary
    print("\n" + "=" * 65)
    print("  AGENT RUN COMPLETE")
    print("=" * 65)
    print(f"\n  📋 Production batches planned  : {len(schedule_df)}")
    print(f"  ✅ On schedule                 : "
          f"{(schedule_df['Status']=='ON_SCHEDULE').sum()}")
    print(f"  ⚠️  At risk of being late       : "
          f"{(schedule_df['Status']=='AT_RISK_LATE').sum()}")
    print(f"  ❌ No tank assigned             : "
          f"{(schedule_df['Status']=='NO_TANK_AVAILABLE').sum()}")
    print(f"  🔔 Alerts generated            : {len(insights)}")
    print(f"\n  📁 Output saved to: {filepath}")
    print(f"  📑 New sheets: Dashboard, Bottleneck_Report, Ingredient_Alerts")
    print("\n" + "=" * 65)

    return schedule_df, insights, material_alerts


if __name__ == "__main__":
    run_agent(EXCEL_FILE)
