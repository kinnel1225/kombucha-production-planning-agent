"""
Build the Kombucha Brewery Excel workbook with all sheets,
sample data, and professional formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date
import os

OUTPUT = "kombucha_brewery.xlsx"
TODAY  = date.today()

# ── Colour palette (no # prefix)
C_NAVY   = "1B3A5C"
C_BLUE   = "2E75B6"
C_LTBLUE = "D6E4F0"
C_GREEN  = "E2EFDA"
C_ORANGE = "FCE4D6"
C_YELLOW = "FFF2CC"
C_RED    = "FCE4D6"
C_WHITE  = "FFFFFF"
C_ALT    = "F2F7FB"
C_BORDER = "B8CCE4"

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header(cell, bg=C_NAVY, size=10):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border    = thin_border()

def data_cell(cell, alt=False, bold=False, align="left"):
    cell.fill      = PatternFill("solid", fgColor=C_ALT if alt else C_WHITE)
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = thin_border()

def autofit(ws, min_w=10, max_w=35):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        mx  = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[ltr].width = max(min_w, min(max_w, mx + 3))

def write_sheet(ws, title_text, col_headers, rows, bg=C_NAVY):
    """Generic sheet builder."""
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(col_headers))}1")
    t = ws["A1"]
    t.value     = title_text
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=bg)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Header row
    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        header(cell, bg=C_BLUE)
    ws.row_dimensions[2].height = 24

    # Data rows
    for ri, row in enumerate(rows, start=3):
        alt = ri % 2 == 0
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            data_cell(c, alt=alt)
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A3"
    autofit(ws)


# ─────────────────────────────────────────────────────────────────
#  BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── 1. ORDERS ────────────────────────────────────────────────────
ws1 = wb.create_sheet("Orders")
order_headers = [
    "Order_ID", "Flavor", "Quantity_Bottles",
    "Due_Date", "Priority", "Customer"
]
d = TODAY
orders_data = [
    ["ORD-001", "Original Kombucha",  1500, str(d + timedelta(days=14)), 1, "Whole Foods Chicago"],
    ["ORD-002", "Ginger Lemon",        800, str(d + timedelta(days=12)), 2, "Fresh Market Peoria"],
    ["ORD-003", "Blueberry Lavender",  600, str(d + timedelta(days=18)), 3, "Natural Grocers"],
    ["ORD-004", "Raspberry Rose",      400, str(d + timedelta(days=20)), 3, "Midwest Co-op"],
    ["ORD-005", "Original Kombucha",  1000, str(d + timedelta(days=21)), 1, "Jewel-Osco"],
    ["ORD-006", "Ginger Lemon",        500, str(d + timedelta(days=25)), 2, "Target Grocery"],
    ["ORD-007", "Mango Turmeric",      700, str(d + timedelta(days=16)), 2, "Fresh Thyme"],
    ["ORD-008", "Blueberry Lavender",  300, str(d + timedelta(days=22)), 3, "Green Earth Market"],
    ["ORD-009", "Original Kombucha",   600, str(d + timedelta(days=30)), 1, "Sam's Club"],
    ["ORD-010", "Mango Turmeric",      400, str(d + timedelta(days=28)), 2, "Sprouts"],
]
write_sheet(ws1, "📋  CUSTOMER ORDERS & DEMAND", order_headers, orders_data)

# ── 2. INVENTORY ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Inventory")
inv_headers = [
    "Material_ID", "Material_Name", "Stock_Level", "Unit",
    "Safety_Stock", "Reorder_Point", "Lead_Time_Days", "Cost_Per_Unit"
]
inv_data = [
    ["MAT-001", "Green Tea (kg)",         45,  "kg",  10,  15,  5,  8.50],
    ["MAT-002", "Black Tea (kg)",          30,  "kg",  8,   12,  5,  6.00],
    ["MAT-003", "Cane Sugar (kg)",         80,  "kg",  20,  30,  3,  1.20],
    ["MAT-004", "SCOBY Starter Culture",   12,  "qty", 3,   5,   7,  25.00],
    ["MAT-005", "Ginger Extract (L)",       8,  "L",   2,   4,   4,  45.00],
    ["MAT-006", "Lemon Juice (L)",         15,  "L",   3,   5,   3,  12.00],
    ["MAT-007", "Blueberry Puree (kg)",     6,  "kg",  2,   4,   6,  18.00],
    ["MAT-008", "Lavender Extract (L)",     3,  "L",   1,   2,   7,  60.00],
    ["MAT-009", "Raspberry Puree (kg)",     9,  "kg",  2,   4,   5,  20.00],
    ["MAT-010", "Rose Water (L)",           4,  "L",   1,   2,   7,  35.00],
    ["MAT-011", "Mango Puree (kg)",        10,  "kg",  3,   5,   5,  15.00],
    ["MAT-012", "Turmeric Powder (kg)",     2,  "kg",  0.5, 1,   4,  22.00],
    ["MAT-013", "Glass Bottles (500mL)",  4800, "qty", 500, 1000, 7,  0.45],
    ["MAT-014", "Bottle Caps",           5000, "qty", 500, 1000, 5,  0.02],
    ["MAT-015", "Labels",               4500, "qty", 500, 800,  4,  0.08],
    ["MAT-016", "Filtered Water (L)",    5000, "L",   500, 800,  1,  0.005],
]
write_sheet(ws2, "🧪  INGREDIENT & PACKAGING INVENTORY", inv_headers, inv_data)

# Low stock highlights
for row_idx in range(3, 3 + len(inv_data)):
    stock_cell   = ws2.cell(row=row_idx, column=3)
    reorder_cell = ws2.cell(row=row_idx, column=6)
    try:
        if stock_cell.value <= reorder_cell.value:
            for ci in range(1, len(inv_headers) + 1):
                ws2.cell(row=row_idx, column=ci).fill = PatternFill(
                    "solid", fgColor="FFE599")
    except Exception:
        pass

# ── 3. RECIPE BOM ─────────────────────────────────────────────────
ws3 = wb.create_sheet("Recipe_BOM")
bom_headers = ["Flavor", "Ingredient", "Quantity_Per_Batch", "Unit", "Notes"]
bom_data = [
    # Original Kombucha (per 200L batch)
    ["Original Kombucha", "Green Tea (kg)",        2.0,  "kg",  "First fermentation base"],
    ["Original Kombucha", "Cane Sugar (kg)",        16.0, "kg",  "7–8% sugar by volume"],
    ["Original Kombucha", "SCOBY Starter Culture",  1.0,  "qty", "1 SCOBY per 200L batch"],
    ["Original Kombucha", "Filtered Water (L)",    180.0, "L",   "Reverse osmosis"],
    ["Original Kombucha", "Glass Bottles (500mL)",  400,  "qty", "200L ÷ 0.5L bottle"],
    ["Original Kombucha", "Bottle Caps",            400,  "qty", "One per bottle"],
    ["Original Kombucha", "Labels",                 400,  "qty", "Brand labels"],
    # Ginger Lemon (per 200L batch)
    ["Ginger Lemon", "Green Tea (kg)",             2.0,  "kg",  "Base tea"],
    ["Ginger Lemon", "Cane Sugar (kg)",            15.0, "kg",  "Slightly less sweet"],
    ["Ginger Lemon", "SCOBY Starter Culture",       1.0, "qty", "1 SCOBY per batch"],
    ["Ginger Lemon", "Ginger Extract (L)",          4.0, "L",   "2F flavor addition"],
    ["Ginger Lemon", "Lemon Juice (L)",             6.0, "L",   "Fresh pressed"],
    ["Ginger Lemon", "Filtered Water (L)",        180.0, "L",   ""],
    ["Ginger Lemon", "Glass Bottles (500mL)",      400,  "qty", ""],
    ["Ginger Lemon", "Bottle Caps",                400,  "qty", ""],
    ["Ginger Lemon", "Labels",                     400,  "qty", ""],
    # Blueberry Lavender
    ["Blueberry Lavender", "Black Tea (kg)",        2.0, "kg",  "Richer base"],
    ["Blueberry Lavender", "Cane Sugar (kg)",      15.0, "kg",  ""],
    ["Blueberry Lavender", "SCOBY Starter Culture", 1.0, "qty", ""],
    ["Blueberry Lavender", "Blueberry Puree (kg)",  8.0, "kg",  "2F addition"],
    ["Blueberry Lavender", "Lavender Extract (L)",  1.5, "L",   "2F addition"],
    ["Blueberry Lavender", "Filtered Water (L)",  175.0, "L",   ""],
    ["Blueberry Lavender", "Glass Bottles (500mL)", 400, "qty", ""],
    ["Blueberry Lavender", "Bottle Caps",           400, "qty", ""],
    ["Blueberry Lavender", "Labels",                400, "qty", ""],
    # Raspberry Rose
    ["Raspberry Rose", "Black Tea (kg)",            2.0, "kg",  ""],
    ["Raspberry Rose", "Cane Sugar (kg)",          14.0, "kg",  ""],
    ["Raspberry Rose", "SCOBY Starter Culture",     1.0, "qty", ""],
    ["Raspberry Rose", "Raspberry Puree (kg)",      7.0, "kg",  "2F addition"],
    ["Raspberry Rose", "Rose Water (L)",            2.0, "L",   "2F addition"],
    ["Raspberry Rose", "Filtered Water (L)",      175.0, "L",   ""],
    ["Raspberry Rose", "Glass Bottles (500mL)",    400,  "qty", ""],
    ["Raspberry Rose", "Bottle Caps",              400,  "qty", ""],
    ["Raspberry Rose", "Labels",                   400,  "qty", ""],
    # Mango Turmeric
    ["Mango Turmeric", "Green Tea (kg)",            2.0, "kg",  ""],
    ["Mango Turmeric", "Cane Sugar (kg)",          14.0, "kg",  ""],
    ["Mango Turmeric", "SCOBY Starter Culture",     1.0, "qty", ""],
    ["Mango Turmeric", "Mango Puree (kg)",          9.0, "kg",  "Alphonso mango"],
    ["Mango Turmeric", "Turmeric Powder (kg)",      0.5, "kg",  "Anti-inflammatory"],
    ["Mango Turmeric", "Filtered Water (L)",      175.0, "L",   ""],
    ["Mango Turmeric", "Glass Bottles (500mL)",    400,  "qty", ""],
    ["Mango Turmeric", "Bottle Caps",              400,  "qty", ""],
    ["Mango Turmeric", "Labels",                   400,  "qty", ""],
]
write_sheet(ws3, "📐  RECIPE / BILL OF MATERIALS (BOM)", bom_headers, bom_data)

# Group by flavor with alternating shade
flavor_colors = {
    "Original Kombucha": "D9EAD3",
    "Ginger Lemon":      "FCE5CD",
    "Blueberry Lavender":"CFE2F3",
    "Raspberry Rose":    "F4CCCC",
    "Mango Turmeric":    "FFF2CC",
}
for row_idx in range(3, 3 + len(bom_data)):
    flavor = ws3.cell(row=row_idx, column=1).value
    if flavor in flavor_colors:
        for ci in range(1, 6):
            ws3.cell(row=row_idx, column=ci).fill = PatternFill(
                "solid", fgColor=flavor_colors[flavor])

# ── 4. FERMENTATION TANKS ─────────────────────────────────────────
ws4 = wb.create_sheet("Fermentation_Tanks")
tank_headers = [
    "Tank_ID", "Capacity_Liters", "Status", "Available_Date",
    "Material", "Temperature_Control", "Notes"
]
tank_data = [
    ["T-01", 200, "Available",  str(TODAY),                 "Stainless Steel 304", "Yes", "Primary fermentation"],
    ["T-02", 200, "Available",  str(TODAY),                 "Stainless Steel 304", "Yes", "Primary fermentation"],
    ["T-03", 200, "In Use",     str(TODAY + timedelta(3)),  "Stainless Steel 304", "Yes", "Currently fermenting Original"],
    ["T-04", 150, "Available",  str(TODAY),                 "Stainless Steel 316", "Yes", "Small batch specialty"],
    ["T-05", 150, "In Use",     str(TODAY + timedelta(5)),  "Stainless Steel 316", "Yes", "Currently fermenting Ginger"],
    ["T-06", 100, "Maintenance",str(TODAY + timedelta(7)),  "Polyethylene HDPE",   "No",  "Scheduled maintenance until +7d"],
    ["T-07", 200, "Available",  str(TODAY),                 "Stainless Steel 304", "Yes", "Secondary/flavor tank"],
    ["T-08", 200, "Available",  str(TODAY),                 "Stainless Steel 304", "Yes", "Secondary/flavor tank"],
]
write_sheet(ws4, "🏭  FERMENTATION TANK INVENTORY", tank_headers, tank_data)

# Color-code status
status_colors = {"Available": "D9EAD3", "In Use": "FCE5CD", "Maintenance": "F4CCCC"}
for row_idx in range(3, 3 + len(tank_data)):
    status_val = ws4.cell(row=row_idx, column=3).value
    color = status_colors.get(status_val, "FFFFFF")
    ws4.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=color)

# ── 5. BOTTLING LINE ─────────────────────────────────────────────
ws5 = wb.create_sheet("Bottling_Line")
btl_headers = [
    "Line_ID", "Line_Name", "Bottles_Per_Hour",
    "Operating_Hours_Per_Day", "Daily_Capacity", "Status", "Notes"
]
btl_data = [
    ["BL-01", "Main Bottling Line",    300, 8, "=C3*D3", "Active",  "Automated rinse-fill-cap"],
    ["BL-02", "Secondary Line",        150, 6, "=C4*D4", "Active",  "Semi-automated"],
    ["BL-03", "Label Application",    500, 8, "=C5*D5", "Active",  "Automated label applicator"],
]
write_sheet(ws5, "🍾  BOTTLING LINE CAPACITY", btl_headers, btl_data)

# ── 6. PRODUCTION PLAN (template) ────────────────────────────────
ws6 = wb.create_sheet("Production_Plan")
plan_headers = [
    "Date", "Flavor", "Batch_Size_Liters", "Assigned_Tank",
    "Fermentation_Start", "Fermentation_End", "Bottling_Date",
    "Actual_Bottling_Date", "Bottles_Planned", "Due_Date", "Priority", "Status"
]
note_row = [
    str(TODAY), "— Run kombucha_agent.py to populate this sheet —",
    "", "", "", "", "", "", "", "", "", ""
]
write_sheet(ws6, "📅  PRODUCTION PLAN (Auto-Generated)", plan_headers, [note_row])

# ── 7. ACTUAL PRODUCTION (log template) ──────────────────────────
ws7 = wb.create_sheet("Actual_Production")
actual_headers = [
    "Date", "Flavor", "Assigned_Tank", "Produced_Bottles",
    "Waste_Bottles", "Yield_Pct", "Remarks"
]
actual_data = [
    [str(TODAY - timedelta(2)), "Original Kombucha",  "T-03", 390, 10, "=E3/D3*100", "Slight over-carbonation on 10 bottles"],
    [str(TODAY - timedelta(3)), "Ginger Lemon",        "T-05", 395, 5,  "=E4/D4*100", "Production within spec"],
    [str(TODAY - timedelta(5)), "Blueberry Lavender",  "T-01", 380, 20, "=E5/D5*100", "Seal issue on line BL-02 — corrected"],
]
write_sheet(ws7, "📊  ACTUAL PRODUCTION LOG", actual_headers, actual_data)

# ── 8. INSTRUCTIONS sheet ────────────────────────────────────────
ws_info = wb.create_sheet("README")
ws_info.sheet_view.showGridLines = False
ws_info.column_dimensions["A"].width = 80

lines = [
    ("KOMBUCHA BREWERY — PRODUCTION PLANNING AI AGENT", C_NAVY, 14, True),
    ("Industrial Engineering Portfolio Project", C_BLUE, 11, False),
    ("", None, 10, False),
    ("HOW TO USE THIS WORKBOOK", C_NAVY, 11, True),
    ("1. Review and update the 'Orders' sheet with current customer demand.", None, 10, False),
    ("2. Update 'Inventory' with current stock levels.", None, 10, False),
    ("3. Check 'Fermentation_Tanks' — update Status and Available_Date.", None, 10, False),
    ("4. Run the Python agent:  python kombucha_agent.py", None, 10, False),
    ("5. The agent will populate: Production_Plan, Dashboard, Bottleneck_Report, Ingredient_Alerts.", None, 10, False),
    ("", None, 10, False),
    ("SHEETS OVERVIEW", C_NAVY, 11, True),
    ("  Orders              → Customer demand and due dates", None, 10, False),
    ("  Inventory           → Raw material and packaging stock", None, 10, False),
    ("  Recipe_BOM          → Bill of materials per flavor", None, 10, False),
    ("  Fermentation_Tanks  → Tank capacity and availability", None, 10, False),
    ("  Bottling_Line       → Line capacity constraints", None, 10, False),
    ("  Production_Plan     → AUTO-GENERATED brewing schedule", None, 10, False),
    ("  Actual_Production   → Manual production log", None, 10, False),
    ("  Dashboard           → AUTO-GENERATED KPI dashboard", None, 10, False),
    ("  Bottleneck_Report   → AUTO-GENERATED insights & alerts", None, 10, False),
    ("  Ingredient_Alerts   → AUTO-GENERATED shortage alerts", None, 10, False),
    ("", None, 10, False),
    ("KEY PARAMETERS (edit in kombucha_agent.py)", C_NAVY, 11, True),
    ("  LITERS_PER_BOTTLE  = 0.355  (12 oz standard bottle)", None, 10, False),
    ("  FERMENTATION_DAYS  = 10     (adjust for your process)", None, 10, False),
    ("  PLANNING_HORIZON   = 21     (days to plan ahead)", None, 10, False),
]

for r_idx, (text, color, size, bold) in enumerate(lines, start=2):
    cell = ws_info.cell(row=r_idx, column=1, value=text)
    fill_color = color if color else C_WHITE
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(name="Arial", size=size, bold=bold,
                     color="FFFFFF" if color else "333333")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws_info.row_dimensions[r_idx].height = 20

# ─────────────────────────────────────────────────────────────────
#  SAVE
# ─────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"✔  Workbook saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
